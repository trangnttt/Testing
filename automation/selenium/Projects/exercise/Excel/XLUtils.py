import openpyxl
from openpyxl.styles import PatternFill

def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(filename=file)
    sheet = workbook[sheetName]
    return(sheet.max_row)

def getColumnCount(file, sheetName):
    workbook = openpyxl.load_workbook(filename=file)
    sheet = workbook[sheetName]
    return(sheet.max_column)

def readData(file, sheetName, row_num, col_num):
    workbook = openpyxl.load_workbook(filename=file)
    sheet = workbook[sheetName]
    return sheet.cell(row_num, col_num).value

def writeData(file, sheetName, row_num, col_num, data):
    workbook = openpyxl.load_workbook(filename=file)
    sheet = workbook[sheetName]
    sheet.cell(row_num, col_num).value = data
    workbook.save(file)

def fillGreenColor(file, sheetName, row_num, col_num):
    workbook = openpyxl.load_workbook(filename=file)
    sheet = workbook[sheetName]
    greenFill = PatternFill(start_color='60b212',
                            end_color='60b212',
                            fill_type='solid')
    sheet.cell(row_num, col_num).fill = greenFill
    workbook.save(file)

def fillRedColor(file, sheetName, row_num, col_num):
    workbook = openpyxl.load_workbook(filename=file)
    sheet = workbook[sheetName]
    redFill = PatternFill(start_color='ff0000',
                          end_color='ff0000',
                          fill_type='solid')
    sheet.cell(row_num, col_num).fill = redFill
    workbook.save(file)