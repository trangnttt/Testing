from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
import time

# Lấy dataTest từ file excel
wb = load_workbook(filename="demo-excel.xlsx")
sh = wb["Sheet"]
total_rows = sh.max_row
total_column = sh.max_column


for elm in range(2, total_rows+1):
    stt = sh.cell(row=elm, column=1).value
    username = sh.cell(row=elm, column=2).value
    password = sh.cell(row=elm, column=3).value
    message = sh.cell(row=elm, column=5).value

    if username is None: 
        username = ""
    if password is None: 
        password = ""
        
    drive = webdriver.Chrome()
    drive.get("https://cms.anhtester.com/login")
    elm1 = drive.find_element(By.ID, "email").send_keys(username)
    elm2 = drive.find_element(By.ID, "password").send_keys(password)
    drive.find_element(By.XPATH, '//button[@class="btn btn-primary btn-lg btn-block"]').click()

    print("TC",stt, message)

time.sleep(4)