
from openpyxl import Workbook, load_workbook

wb = load_workbook(filename="demo-excel.xlsx")
sh = wb["Sheet"]

print(sh["A3"].value)
print(wb["Sheet"]["A2"].value)

print(sh.cell(row=3, column=5).value)