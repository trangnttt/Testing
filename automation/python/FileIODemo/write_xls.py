# Cài: pip3 install openpyxl

from openpyxl import Workbook
wb = Workbook()

ws = wb.active
ws['A1'] = "Hello nha"
datas = (
    ['ankit', 1000],
    ['rahul',   100],
    ['priya',  300],
    ['harshita',    50],
)
for data in datas:
    ws.append(data)

for i in range(4, 6):
    for j in range(3, 6):
        ws.cell(row=i, column = j).value = i + j
wb.save("demo-excel.xlsx")